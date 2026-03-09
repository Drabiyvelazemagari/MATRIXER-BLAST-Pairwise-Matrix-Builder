#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
BlastMatrixApp_v11.pyw

Hierarchy mode removed.
Single-file BLAST input only.

Matrix behavior identical to original engine.

Rule added:
    values <= 1 → "≤1"z
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def format_value(v):
    try:
        f = float(v)
        if f <= 1:
            return "≤1"
        return f
    except:
        return v


SUBJECT_HEADER_ROW = 2
SUBJECT_FIRST_COL = 4
QUERY_FIRST_COVER_ROW = 3
QUERY_ID_COL = 2
LABEL_COL = 3
ROWS_PER_QUERY = 2

EMPTY_ONLY = {None, ""}


@dataclass
class Hit:
    accession: str
    query_cover: str
    perc_ident: str


def read_text_or_zip(path: Path) -> str:
    if path.suffix.lower() == ".zip":
        with zipfile.ZipFile(path, "r") as z:
            names = z.namelist()
            txts = [n for n in names if n.lower().endswith(".txt")]
            chosen = txts[0] if txts else names[0]
            with z.open(chosen, "r") as f:
                data = f.read()
        return data.decode(errors="replace")
    return path.read_text(errors="replace")


def parse_percent(val: str) -> float:
    s = str(val).strip()
    if s.endswith("%"):
        s = s[:-1].strip()
    return float(s)


def is_accession_version(s: str) -> bool:
    s = s.strip()
    return bool(
        re.fullmatch(r"[A-Z]{1,4}\d+\.\d+", s) or
        re.fullmatch(r"[A-Z]{1,3}_\d+\.\d+", s) or
        re.fullmatch(r"NZ_[A-Z]{1,4}\d+\.\d+", s)
    )


def normalize_query_id(raw: str) -> str:
    raw = raw.strip().strip("|")
    if "|" not in raw:
        return raw
    parts = [p for p in raw.split("|") if p]
    for p in reversed(parts):
        if "." in p and re.search(r"\.\d+$", p):
            return p
    return parts[-1]


def extract_query_id_from_line(line: str) -> Optional[str]:
    s = line.strip()

    m = re.match(r"^Query\s+#\d+:\s.*?\bQuery\s+ID:\s*([^\s]+)", s)
    if m:
        return normalize_query_id(m.group(1))

    if s.startswith("Query:"):
        m2 = re.search(r"\bID:\s*([^\s]+)", s)
        if m2:
            return normalize_query_id(m2.group(1))
        rem = s.replace("Query:", "").strip()
        return rem or None

    if s.startswith("Query="):
        rem = s.replace("Query=", "").strip()
        return normalize_query_id(rem.split()[0]) if rem else None

    return None


def parse_all_queries_from_report(text: str) -> Tuple[Dict[str, List[Hit]], List[str]]:
    lines = text.splitlines()
    out: Dict[str, List[Hit]] = {}
    order: List[str] = []

    current_qid: Optional[str] = None
    synthetic_counter = 0

    def ensure_qid() -> str:
        nonlocal synthetic_counter, current_qid
        if current_qid and current_qid.strip():
            if current_qid not in out:
                out[current_qid] = []
                order.append(current_qid)
            return current_qid
        synthetic_counter += 1
        current_qid = f"Query_{synthetic_counter}"
        out[current_qid] = []
        order.append(current_qid)
        return current_qid

    i = 0
    while i < len(lines):
        s = lines[i].strip()

        q = extract_query_id_from_line(s)
        if q:
            current_qid = q
            if current_qid not in out:
                out[current_qid] = []
                order.append(current_qid)

        if s.startswith("Sequences producing significant alignments"):
            qid = ensure_qid()
            j = i + 1

            while j < len(lines):
                ln = lines[j].strip()
                if ln == "":
                    break

                tokens = ln.split()
                if len(tokens) >= 6:
                    accession = tokens[-1]
                    if is_accession_version(accession):
                        qcov = tokens[-5]
                        pident = tokens[-3]
                        out[qid].append(Hit(accession, qcov, pident))
                j += 1

            i = j
            continue

        i += 1

    return out, order


def read_existing_matrix(ws: Worksheet):
    subjects = []
    col = SUBJECT_FIRST_COL
    while True:
        v = ws.cell(row=SUBJECT_HEADER_ROW, column=col).value
        if v in EMPTY_ONLY:
            break
        subjects.append(str(v))
        col += 1

    queries = []
    row = QUERY_FIRST_COVER_ROW
    while True:
        v = ws.cell(row=row, column=QUERY_ID_COL).value
        if v in EMPTY_ONLY:
            break
        queries.append(str(v))
        row += ROWS_PER_QUERY

    values = {}

    for i, q in enumerate(queries):
        r_cov = QUERY_FIRST_COVER_ROW + i * 2
        r_id = r_cov + 1

        for j, s in enumerate(subjects):
            c = SUBJECT_FIRST_COL + j

            cov_val = ws.cell(row=r_cov, column=c).value
            id_val = ws.cell(row=r_id, column=c).value

            if cov_val not in EMPTY_ONLY:
                values[(q, s, "cov")] = cov_val

            if id_val not in EMPTY_ONLY:
                values[(q, s, "id")] = id_val

    return subjects, queries, values


def build_universe(existing_order, primary_order):
    seen = set()
    out = []

    for x in existing_order + primary_order:
        if x not in seen:
            seen.add(x)
            out.append(x)

    return out


def write_matrix(ws, universe, preserved, blast_pairs):

    ws.cell(row=2, column=2).value = "Accession Number"

    for j, subj in enumerate(universe):
        ws.cell(row=2, column=SUBJECT_FIRST_COL + j).value = subj

    for i, q in enumerate(universe):

        r_cov = QUERY_FIRST_COVER_ROW + i * 2
        r_id = r_cov + 1

        ws.cell(row=r_cov, column=QUERY_ID_COL).value = q
        ws.cell(row=r_cov, column=LABEL_COL).value = "Query Cover (%)"
        ws.cell(row=r_id, column=LABEL_COL).value = "DNA identity (%)"

        for j, s in enumerate(universe):

            c = SUBJECT_FIRST_COL + j

            cov_key = (q, s, "cov")
            id_key = (q, s, "id")

            if q == s:
                ws.cell(row=r_cov, column=c).value = "-"
                ws.cell(row=r_id, column=c).value = "-"
                continue

            if cov_key in preserved:
                ws.cell(row=r_cov, column=c).value = preserved[cov_key]

            elif cov_key in blast_pairs:
                ws.cell(row=r_cov, column=c).value = format_value(blast_pairs[cov_key])

            else:
                ws.cell(row=r_cov, column=c).value = "NSS"

            if id_key in preserved:
                ws.cell(row=r_id, column=c).value = preserved[id_key]

            elif id_key in blast_pairs:
                ws.cell(row=r_id, column=c).value = format_value(blast_pairs[id_key])

            else:
                ws.cell(row=r_id, column=c).value = "NSS"

# =========================
# VISUAL LAYOUT + MERGES
# =========================

THIN = Side(style="thin")
MED = Side(style="medium")

BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_LEFT_MED = Border(left=MED, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM_MED = Border(left=THIN, right=THIN, top=THIN, bottom=MED)
BORDER_LEFT_BOTTOM_MED = Border(left=MED, right=THIN, top=THIN, bottom=MED)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center")

FONT_BOLD = Font(bold=True)


def apply_visual_layout_and_merges(ws, universe):

    n = len(universe)

    max_col = SUBJECT_FIRST_COL + n - 1
    max_row = QUERY_FIRST_COVER_ROW + (n * 2) - 1

    ws.freeze_panes = "D3"

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24

    for col in range(SUBJECT_FIRST_COL, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = 16

    # Merge "Accession Number"
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)

    # Merge query IDs
    for i in range(n):

        r_cov = QUERY_FIRST_COVER_ROW + i * 2
        r_id = r_cov + 1

        ws.merge_cells(
            start_row=r_cov,
            start_column=QUERY_ID_COL,
            end_row=r_id,
            end_column=QUERY_ID_COL,
        )

    # Styling
    for i in range(n):

        r_cov = QUERY_FIRST_COVER_ROW + i * 2
        r_id = r_cov + 1

        ws.cell(row=r_cov, column=QUERY_ID_COL).alignment = ALIGN_CENTER

        cov_label = ws.cell(row=r_cov, column=LABEL_COL)
        id_label = ws.cell(row=r_id, column=LABEL_COL)

        cov_label.font = FONT_BOLD
        id_label.font = FONT_BOLD

        cov_label.alignment = ALIGN_LEFT_CENTER
        id_label.alignment = ALIGN_LEFT_CENTER

        for col in range(SUBJECT_FIRST_COL, max_col + 1):

            c1 = ws.cell(row=r_cov, column=col)
            c2 = ws.cell(row=r_id, column=col)

            c1.alignment = ALIGN_CENTER
            c2.alignment = ALIGN_CENTER

            if col == SUBJECT_FIRST_COL:
                c1.border = BORDER_LEFT_MED
                c2.border = BORDER_LEFT_BOTTOM_MED
            else:
                c1.border = BORDER_THIN
                c2.border = BORDER_BOTTOM_MED

class App(tk.Tk):

    def __init__(self):

        super().__init__()

        self.title("BLAST → Matrix Builder")
        self.geometry("1000x700")

        self.in_single = tk.StringVar()
        self.in_existing_xlsx = tk.StringVar()

        self.out_folder = tk.StringVar()
        self.out_name = tk.StringVar(value=".xlsx")

        self._build()

    def _build(self):

        pad = 6

        tk.Label(self, text="BLAST report (.txt or .zip):").grid(row=0, column=0, padx=pad, pady=pad)
        tk.Entry(self, textvariable=self.in_single, width=90).grid(row=0, column=1, padx=pad, pady=pad)
        tk.Button(self, text="Browse", command=self.browse_single).grid(row=0, column=2)

        tk.Label(self, text="Existing matrix (.xlsx) [optional]:").grid(row=1, column=0, padx=pad, pady=pad)
        tk.Entry(self, textvariable=self.in_existing_xlsx, width=90).grid(row=1, column=1)
        tk.Button(self, text="Browse", command=self.browse_xlsx).grid(row=1, column=2)

        tk.Label(self, text="Output folder:").grid(row=2, column=0, padx=pad, pady=pad)
        tk.Entry(self, textvariable=self.out_folder, width=90).grid(row=2, column=1)
        tk.Button(self, text="Browse", command=self.browse_out_folder).grid(row=2, column=2)

        tk.Label(self, text="Output filename:").grid(row=3, column=0, padx=pad, pady=pad)
        tk.Entry(self, textvariable=self.out_name, width=90).grid(row=3, column=1)

        tk.Button(self, text="Run", command=self.run, height=2).grid(row=4, column=1)

        self.log_box = ScrolledText(self, width=130, height=26)
        self.log_box.grid(row=5, column=0, columnspan=3)

    def browse_single(self):
        p = filedialog.askopenfilename(filetypes=[("BLAST", "*.txt *.zip")])
        if p:
            self.in_single.set(p)

    def browse_xlsx(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.in_existing_xlsx.set(p)

    def browse_out_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.out_folder.set(p)

    def run(self):

        try:

            p = Path(self.in_single.get().strip())
            if not p.exists():
                messagebox.showerror("Error", "Please select one BLAST report.")
                return

            single_map, single_order = parse_all_queries_from_report(read_text_or_zip(p))

            merged = single_map
            primary_order = single_order

            preserved_values = {}
            existing_order = []

            existing_path = self.in_existing_xlsx.get().strip()
            if existing_path:
                old_wb = load_workbook(existing_path)
                old_ws = old_wb.active
                ex_subjects, ex_queries, ex_vals = read_existing_matrix(old_ws)
                preserved_values = ex_vals
                existing_order = ex_subjects or ex_queries

            universe = build_universe(existing_order, primary_order)

            universe_set = set(universe)

            blast_pairs = {}

            for qid, hits in merged.items():

                if qid not in universe_set:
                    continue

                for h in hits:

                    s = h.accession

                    if s not in universe_set or s == qid:
                        continue

                    qcov = round(parse_percent(h.query_cover), 2)
                    pid = round(float(str(h.perc_ident).strip()), 2)

                    blast_pairs[(qid, s, "cov")] = qcov
                    blast_pairs[(qid, s, "id")] = pid

            wb = Workbook()
            ws = wb.active

            write_matrix(ws, universe, preserved_values, blast_pairs)

            apply_visual_layout_and_merges(ws, universe)

            out_dir = Path(self.out_folder.get().strip())
            out_name = self.out_name.get().strip()

            if not out_name.lower().endswith(".xlsx"):
                out_name += ".xlsx"

            out_path = out_dir / out_name

            wb.save(out_path)

            messagebox.showinfo("Done", f"Saved:\n{out_path}")

        except Exception as e:
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    App().mainloop()